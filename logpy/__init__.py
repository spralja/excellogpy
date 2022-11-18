__version__ = '0.1.0'

from logpy.model import Entry, Log

from pathlib import Path
from typing import Union
from datetime import datetime, timedelta

from openpyxl import load_workbook, Workbook


DAYS = {
    'Monday': 1,
    'Tuesday': 2,
    'Wednesday': 3,
    'Thursday': 4,
    'Friday': 5,
    'Saturday': 6,
    'Sunday': 7
}


def load_log(path: Union[Path, str]) -> Log:
    """
    Loading log from folder
    :param path: Union[Path, str]
    :return: Log
    """
    path = Path(path)
    log = Log(path.name)
    for year_dir in path.iterdir():
        year = int(year_dir.name)
        for week_dir in year_dir.iterdir():
            week = int(week_dir.name.split('.')[0][1:])
            work_book = load_workbook(week_dir, read_only=True, data_only=True)
            for day in DAYS.keys():
                date = datetime.fromisocalendar(year, week, DAYS[day])
                try: sheet = work_book[day]
                except KeyError: continue

                for row in sheet.iter_rows(3):
                    row = [cell.value for cell in row]
                    if not row[0]: break

                    start_time = {
                        'hours': row[0].hour,
                        'minutes': row[0].minute
                    }

                    if row[1].hour == 0:
                        end_time = {
                            'days': 1
                        }
                    else:
                        end_time = {
                            'hours': row[1].hour,
                            'minutes': row[1].minute
                        }

                    row[0] = date + timedelta(**start_time)
                    row[1] = date + timedelta(**end_time)

                    log.add_entry(Entry(*row))

            work_book.close()

    return log


def export_log(log: Log, path: Union[Path, str]):
    """
    Exports the log in the give path (does not support entries which span multiple days)
    :param log:
    :param path:
    :return:
    """
    path = Path(path)

    if path.is_dir() or path.is_file():
        raise FileExistsError(f'Directory {path.name} already exists.')

    path.mkdir()

    entries = list(log.range())
    entries.sort()
    wb: Workbook = None
    curr = None
    for entry in entries:
        year = str(entry.start_time.year)
        week = str(entry.start_time.isocalendar().week)
        day = entry.start_time.strftime('%A')
        curr = path / year / f'w{week}.xlsx'
        print(curr)
        if not (path / year).is_dir():
            (path / year).mkdir()

        if not curr.is_file():
            if wb is not None:
                wb.save(curr)
            wb = Workbook()
            del wb['Sheet']

        try: wb[day]
        except KeyError:
            wb.create_sheet(day)
            wb[day].append([])
            wb[day].append([])

        row = [
            entry.start_time,
            entry.end_time,
            entry.category,
            entry.description
        ]

        print(wb)
        wb[day].append(row)
        wb[day][f'A{wb[day].max_row}'].number_format = 'h:mm'
        wb[day][f'B{wb[day].max_row}'].number_format = 'h:mm'

    wb.save(curr)
